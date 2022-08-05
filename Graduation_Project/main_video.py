import time
import cv2
import openpyxl
from facerec import Facerec

wb = openpyxl.load_workbook("detected_face_names.xlsx")
sh1 = wb['sayfa1']
# Encode faces from a folder
sfr = Facerec()
sfr.load_encoding_images("images/")

# Load Camera
cap = cv2.VideoCapture(1)

i = 1
while True:
    none, frame = cap.read()
    # Detect Faces
    face_locations, face_names = sfr.detect_known_faces(frame)
    for face_loc, name in zip(face_locations, face_names):
        y1, x2, y2, x1 = face_loc[0], face_loc[1], face_loc[2], face_loc[3]

        cv2.putText(frame, name,(x1, y1 - 10), cv2.FONT_HERSHEY_DUPLEX, 1, (0, 0, 200), 2)
        cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 200), 5)
        sh1.cell(row=i + 1, column=1, value=name)
        time_now=time.asctime(time.localtime())
        sh1.cell(row=i+1,column=2,value=time_now)
        wb.save("detected_face_names.xlsx")
        i = i + 1
        print(name,time_now)



    cv2.imshow("Frame", frame)

    key = cv2.waitKey(1)
    if key == 1:
        break

cap.release()
cv2.destroyAllWindows()
